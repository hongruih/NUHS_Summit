"""
=============================================================================
  REAL-TIME SENTIMENT ANALYSIS & WORD CLOUD BOOTH TOOL  
  For NUHS Summit — Interactive Audience Engagement
=============================================================================


    ✓ AI vs Human Turing Test module — clinical scenarios with two answers
    ✓ Respondents guess which answer is AI vs Human, then rate trust/empathy/safety/usefulness
    ✓ Demographic capture: job group + seniority level
    ✓ Dashboard analytics: % correct by job group/seniority (tests Dr Alex's hypothesis)
    ✓ QR-code accessible standalone survey page (/survey)
    ✓ Results flash on the main dashboard in real-time
    ✓ SQLite persistence (survives restarts)
    ✓ All previous features preserved (live transcription, word clouds, sentiment)

  Layout:
    Tab 1: LIVE OVERVIEW       — Aggregate word cloud + sentiment + Turing snapshot
    Tab 2: Q1 Input            — Record/type + per-question dashboard
    Tab 3: Q2 Input            — Record/type + per-question dashboard
    Tab 4: Q3 Input            — Record/type + per-question dashboard
    Tab 5: PER-Q DASHBOARD     — All 3 questions word clouds + sentiments
    Tab 6: AI vs HUMAN         — Turing test survey + results dashboard

  Standalone:
    /survey                    — Mobile-friendly QR survey (walk-away)
    /qr                        — QR code image for /survey

  Requirements:
      pip install flask textblob qrcode pillow

  Usage:
      python app.py  →  open http://localhost:5001

  Authors: Alexander Yip, Ravi Shankar, Elya Chen, Emily Chew & HCRD
=============================================================================
"""

import os, io, re, base64, json, sqlite3, uuid
import openpyxl
from openpyxl.styles import Font
from dotenv import load_dotenv
load_dotenv()
from collections import Counter
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file, Response, redirect, url_for
from textblob import TextBlob

# spaCy — optional but recommended for lemmatised word clouds
# Install: pip install spacy && python -m spacy download en_core_web_sm
try:
    import spacy
    _nlp = spacy.load("en_core_web_sm", disable=["parser", "ner"])
    _SPACY_OK = True
except Exception:
    _nlp = None
    _SPACY_OK = False

# Profanity filter — screens submissions and word cloud output
# Install: pip install better-profanity
from better_profanity import profanity as _profanity
_profanity.load_censor_words()

APP_TITLE = "AI Perspectives — Live Sentiment Booth"
QUESTIONS = [
    "What changes, if any, do you anticipate AI bringing to your work?",
    "What would make you more confident or willing to use AI tools in your daily practice?",
    "What concerns, if any, do you have about the use of AI in healthcare?",
    "How do you think AI could best support, rather than replace, the human elements of healthcare?",
]
STOP_WORDS = set("i me my myself we our ours ourselves you your yours yourself yourselves "
    "he him his himself she her hers herself it its itself they them their theirs "
    "themselves what which who whom this that these those am is are was were be been "
    "being have has had having do does did doing a an the and but if or because as "
    "until while of at by for with about against between through during before after "
    "above below to from up down in out on off over under again further then once "
    "here there when where why how all both each few more most other some such no nor "
    "not only own same so than too very s t can will just don should now d ll m o re "
    "ve y ain aren couldn didn doesn hadn hasn haven isn ma mightn mustn needn shan "
    "shouldn wasn weren won wouldn think also like really would could one thing things "
    "going know actually mean yeah yes well okay sure right get got say said something "
    "lot much many still even want need gonna make way feel see look come kind sort "
    "maybe might quite already".split())

# ═══════════════════════════════════════════════════════
# AI vs HUMAN TURING TEST — ITEMS
# Section 1: 3 Scenarios (single response, judge human vs AI)
# Section 2: 2 Conversations (multi-turn dialog, judge the clinician)
# "is_ai" / "correct_answer" are NEVER sent to the survey frontend;
#   they are embedded in the client JS for inline reveal (booth activity, no security concern)
# ═══════════════════════════════════════════════════════
TURING_ITEMS = [
    # ─── Section 1: Scenarios ────────────────────────────────────────────────
    {
        "id": "scenario_1",
        "section": "Scenario",
        "label": "Scenario 1",
        "type": "scenario",
        "patient": (
            "I've tried to lose weight many times, but I always gain it back. "
            "I feel like I'm failing, and maybe this is just who I am. What should I do?"
        ),
        "response": (
            "I can see your effort towards making sustained weight loss and it sounds like "
            "this cycle has impacted your own self perception. On the other hand, it seems "
            "that you also have had success in losing weight despite still figuring out ways "
            "to ensure it is maintained."
        ),
        "is_ai": False,
        "correct_answer": "Human Clinician",
    },
    {
        "id": "scenario_2",
        "section": "Scenario",
        "label": "Scenario 2",
        "type": "scenario",
        "patient": (
            "I've been feeling really down for the past few months. I don't enjoy things "
            "I used to, and I'm sleeping a lot more. Is this just a phase?"
        ),
        "response": (
            "Thank you for sharing. Your symptoms — persistent low mood, anhedonia, and "
            "hypersomnia lasting several months — align with diagnostic criteria for a "
            "depressive episode (DSM-5) and are unlikely to represent a transient phase.\n"
            "Recommended actions:\n"
            "• Consult a licensed mental health professional or your GP for a formal assessment\n"
            "• A PHQ-9 screening may be administered to quantify symptom severity\n"
            "• Evidence-based treatment options include CBT, pharmacotherapy, or a combination"
        ),
        "is_ai": True,
        "correct_answer": "AI",
    },
    {
        "id": "scenario_3",
        "section": "Scenario",
        "label": "Scenario 3",
        "type": "scenario",
        "patient": (
            "My child has had a fever of 38.5\u00b0C for two days and a rash appeared this morning. "
            "She's still eating and playing. Can I just give paracetamol and wait?"
        ),
        "response": (
            "Thank you for your question. A concurrent fever (48h) and new-onset rash requires "
            "same-day medical evaluation, even if the child appears well. Possible differentials "
            "include viral exanthem, scarlet fever, or \u2014 less commonly \u2014 meningococcal disease "
            "or Kawasaki disease.\n"
            "Paracetamol may be given at the appropriate weight-based dose in the interim. "
            "Seek emergency care immediately if:\n"
            "\u2022 Rash is non-blanching\n"
            "\u2022 Child becomes lethargic or refuses fluids\n"
            "\u2022 Fever exceeds 40\u00b0C\n"
            "\u2022 Neck stiffness or photophobia develops"
        ),
        "is_ai": True,
        "correct_answer": "AI",
    },
    # ─── Section 2: Conversations ─────────────────────────────────────────────
    {
        "id": "conversation_1",
        "section": "Conversation",
        "label": "Conversation 1",
        "type": "conversation",
        "turns": [
            {"speaker": "Patient", "text": "Hi Doctor, I feel a tightness in my chest when I walk upstairs. It goes away when I rest. Should I just monitor it at home?"},
            {"speaker": "Doctor",  "text": "How often do you feel this tightness?"},
            {"speaker": "Patient", "text": "Every day when I walk upstairs"},
            {"speaker": "Doctor",  "text": "Do you experience any other symptoms?"},
            {"speaker": "Patient", "text": "No leh, just tightness in the chest"},
            {"speaker": "Doctor",  "text": "Did you perform any strenuous exercise before the symptoms"},
            {"speaker": "Patient", "text": "I don't think so..."},
            {"speaker": "Doctor",  "text": "Anyone in the family has heart conditions?"},
            {"speaker": "Patient", "text": "My grandfather had a heart attack many years back"},
        ],
        "is_ai": False,
        "correct_answer": "Human Clinician",
    },
    {
        "id": "conversation_2",
        "section": "Conversation",
        "label": "Conversation 2",
        "type": "conversation",
        "turns": [
            {"speaker": "Patient",    "text": "I accidentally took double my blood pressure medication this morning. I feel okay so far. Should I be worried?"},
            {"speaker": "Pharmacist", "text": "Thank you for flagging this. To assess the clinical risk accurately, could you confirm your currently prescribed dose and medication name?"},
            {"speaker": "Patient",    "text": "I'm not sure leh"},
            {"speaker": "Pharmacist", "text": "Understood. To assist with medication identification, please describe the shape of the tablet."},
            {"speaker": "Patient",    "text": "Oval shape"},
            {"speaker": "Pharmacist", "text": "Noted. Please specify the tablet colour"},
            {"speaker": "Patient",    "text": "White"},
            {"speaker": "Pharmacist", "text": (
                "Based on the parameters provided \u2014 oval, white, approximately 8\u201310mm \u2014 this is "
                "consistent with several commonly prescribed antihypertensives, including amlodipine "
                "5mg or perindopril 4mg. A double dose of most first-line antihypertensives carries "
                "a risk of hypotension, reflex tachycardia or bradycardia, and peripheral oedema. "
                "As you are currently asymptomatic, monitor closely for the next 4\u20136 hours. Avoid "
                "strenuous activity. If you experience syncope, chest discomfort, or sustained "
                "dizziness, proceed to the nearest A&E immediately."
            )},
        ],
        "is_ai": True,
        "correct_answer": "AI",
    },
]

JOB_GROUPS = ["Junior Doctor / HO / MO", "Senior Doctor / Consultant", "Nurse", "Allied Health Professional",
              "Clinical Research Coordinator / Scientist", "Administrator / Management", "Student", "Other"]
SENIORITY_LEVELS = ["Junior (< 5 years experience)", "Mid-career (5–15 years)", "Senior (> 15 years)"]
TRUST_TASKS = ["Triaging symptoms", "Drafting clinical notes", "Patient education",
               "Medication counselling", "Mental health screening", "Diagnostic support"]

# ═══════════════════════════════════════════════════════
# AI ACCEPTANCE SURVEY — CONFIG
# ═══════════════════════════════════════════════════════
ACCEPTANCE_PART_A = {
    "age_group":        ["Under 30", "30–39", "40–49", "50–59", "60 or above"],
    "gender":           ["Male", "Female"],
    "cluster":          ["National University Health System", "National Healthcare Group",
                         "SingHealth", "Others"],
    "disciplines":      ["Medicine", "Nursing", "Allied Health Professions", "Healthcare Research",
                         "Healthcare Administration / Operations", "Other"],
    "years_healthcare": ["Less than 1 year", "1–5 years", "6–10 years", "11–20 years", "More than 20 years"],
    "years_role":       ["Less than 1 year", "1–3 years", "4–7 years", "8–15 years", "More than 15 years"],
    "ai_frequency":     ["Always", "Often", "Sometimes", "Rarely", "Never"],
    "ai_tools":         ["Commercial AI (ChatGPT, Gemini, Claude)",
                         "Institutional AI (Medivoice, RussellGPT, BotNUHS, Notebuddy, etc.)",
                         "Government AI (Pair, Transcribe, Tandem)",
                         "Clinically Integrated AI (tools supporting clinical decision-making, diagnosis, e.g. Automated Visual Acuity Test)"],
}

# Likert questions per part. Keys match the JSON keys stored in the DB (e.g. "B1", "C3").
# Questions containing [discipline] are substituted client-side based on the respondent's discipline.
ACCEPTANCE_LIKERT = {
    "B": {
        "title": "AI, Deskilling, and Upskilling",
        "questions": {
            "B1": "I believe that relying on AI tools causes me to lose core [discipline] skills over time.",
            "B2": "I believe that relying on AI tools causes me to lose professional communication skills over time.",
            "B3": "I believe that overrelying on AI recommendations reduces the quality of independent [discipline] reasoning.",
            "B4": "I believe that early-career staff may not develop foundational competencies if AI handles too many tasks for them.",
            "B5": "I believe that with the help of AI, early-career staff may miss opportunities to grapple with complex issues that foster adaptive learning.",
            "B6": "I believe that AI tools free up my time to develop higher-order [discipline] skills.",
            "B7": "I believe that reviewing AI recommendations actively sharpens my [discipline] skills.",
        }
    },
    "C": {
        "title": "Perception of AI Adoption",
        "questions": {
            "C1": "I believe that AI makes my specialised [discipline] knowledge appear less valuable to others.",
            "C2": "I believe that my professional standing within healthcare is undermined if AI can perform the tasks I was trained for.",
            "C3": "I would not want AI to change the way I currently work.",
            "C4": "I would not rely on AI recommendations to make [discipline] decisions.",
            "C5": "I would not feel comfortable using AI in high-stakes decisions in my work.",
            "C6": "I believe that the benefits of AI in healthcare outweigh the risks or costs of implementing it.",
            "C7": "I believe that AI tools improve the quality of decisions in my organisation.",
            "C8": "I believe that AI tools reduce my administrative burden.",
        }
    },
    "D": {
        "title": "Organisational Culture and Support for AI",
        "questions": {
            "D1": "I believe that my organisation actively promotes the use of AI at work.",
            "D2": "I believe that it is safe to raise concerns or questions about AI in my organisation.",
            "D3": "I believe that my organisation considers employees' feedback in the selection and deployment of AI tools.",
            "D4": "I believe that my organisation prioritises AI implementation.",
            "D5": "I believe that my organisation provides enough resources and training to support the use of AI.",
            "D6": "I believe that I am able to use AI tools confidently in my role.",
            "D7": "I believe that I am capable of critically evaluating and overriding AI recommendations when necessary.",
        }
    },
}

# ═══════════════════════════════════════════════════════
# DATABASE SETUP
# ═══════════════════════════════════════════════════════
DB_PATH = os.environ.get("DATABASE_PATH", "booth_data.db")

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    c = conn.cursor()
    # Sentiment responses (migrated from in-memory store)
    c.execute("""CREATE TABLE IF NOT EXISTS sentiment_responses (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        question_index INTEGER NOT NULL,
        text TEXT NOT NULL,
        polarity REAL,
        subjectivity REAL,
        label TEXT,
        emoji TEXT,
        color TEXT,
        timestamp TEXT
    )""")
    # Turing test responses
    c.execute("""CREATE TABLE IF NOT EXISTS turing_responses (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        respondent_id TEXT NOT NULL,
        job_group TEXT,
        seniority TEXT,
        timestamp TEXT
    )""")
    c.execute("""CREATE TABLE IF NOT EXISTS turing_answers (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        respondent_id TEXT NOT NULL,
        item_id TEXT NOT NULL,
        section TEXT NOT NULL,
        guess TEXT NOT NULL,
        correct INTEGER NOT NULL,
        rating_trust INTEGER,
        rating_empathy INTEGER,
        rating_safety INTEGER,
        rating_usefulness INTEGER,
        timestamp TEXT
    )""")
    # Schema migration: drop and recreate turing_answers if it still has the old schema
    # (old schema used guessed_ai_index + scenario_id; new schema uses item_id + section + guess).
    # Safe to do before the event; after real data exists the table will already have item_id.
    ta_cols = {row[1] for row in c.execute("PRAGMA table_info(turing_answers)").fetchall()}
    if "item_id" not in ta_cols:
        c.execute("DROP TABLE turing_answers")
        c.execute("""CREATE TABLE turing_answers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            respondent_id TEXT NOT NULL,
            item_id TEXT NOT NULL,
            section TEXT NOT NULL,
            guess TEXT NOT NULL,
            correct INTEGER NOT NULL,
            rating_trust INTEGER,
            rating_empathy INTEGER,
            rating_safety INTEGER,
            rating_usefulness INTEGER,
            timestamp TEXT
        )""")
        conn.commit()
    # AI trust tasks
    c.execute("""CREATE TABLE IF NOT EXISTS turing_tasks (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        respondent_id TEXT NOT NULL,
        task TEXT NOT NULL
    )""")
    # AI Acceptance Survey responses
    c.execute("""CREATE TABLE IF NOT EXISTS acceptance_responses (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        participant_id TEXT NOT NULL,
        timestamp TEXT NOT NULL,
        age_group TEXT,
        gender TEXT,
        cluster TEXT,
        disciplines TEXT,
        years_healthcare TEXT,
        years_role TEXT,
        seniority TEXT,
        ai_frequency TEXT,
        ai_tools TEXT,
        likert_answers TEXT,
        open_reflection TEXT
    )""")
    # Migration: add participant_id to sentiment_responses (safe no-op if already exists)
    try:
        c.execute("ALTER TABLE sentiment_responses ADD COLUMN participant_id TEXT")
        conn.commit()
    except sqlite3.OperationalError:
        pass
    # Migration: add cluster to acceptance_responses
    try:
        c.execute("ALTER TABLE acceptance_responses ADD COLUMN cluster TEXT")
        conn.commit()
    except sqlite3.OperationalError:
        pass
    # Migration: add open_reflection to acceptance_responses
    try:
        c.execute("ALTER TABLE acceptance_responses ADD COLUMN open_reflection TEXT")
        conn.commit()
    except sqlite3.OperationalError:
        pass
    conn.commit()
    conn.close()

init_db()

app = Flask(__name__)

# ═══════════════════════════════════════════════════════
# SENTIMENT HELPERS (unchanged logic, now with SQLite)
# ═══════════════════════════════════════════════════════
def sentiment(text):
    b = TextBlob(text)
    p = b.sentiment.polarity
    s = b.sentiment.subjectivity
    if p > 0.15:
        return {"polarity": round(p, 3), "subjectivity": round(s, 3),
                "label": "Positive", "emoji": "😊", "color": "#22c55e"}
    elif p < -0.15:
        return {"polarity": round(p, 3), "subjectivity": round(s, 3),
                "label": "Negative", "emoji": "😟", "color": "#ef4444"}
    else:
        return {"polarity": round(p, 3), "subjectivity": round(s, 3),
                "label": "Neutral", "emoji": "😐", "color": "#f59e0b"}

def extract_words(texts):
    """
    Return [{text, value}] sorted by frequency, capped at 80 words.

    When spaCy is available (en_core_web_sm):
      - Lemmatises tokens  (concerns/concerned/concerning → concern)
      - Keeps only NOUN, VERB, ADJ tokens
      - Removes stop words via spaCy's built-in list merged with STOP_WORDS
    Falls back to the original Counter approach if spaCy is unavailable.

    Install spaCy:
        pip install spacy
        python -m spacy download en_core_web_sm
    Add 'spacy' to requirements.txt after installing.
    """
    combined = " ".join(texts)

    if _SPACY_OK:
        # Merge the project's curated stop words into spaCy's set so nothing is lost
        spacy_stops = _nlp.Defaults.stop_words | STOP_WORDS

        doc = _nlp(combined)
        lemmas = [
            token.lemma_.lower()
            for token in doc
            if token.pos_ in {"NOUN", "VERB", "ADJ"}
            and token.is_alpha
            and len(token.lemma_) > 2
            and token.lemma_.lower() not in spacy_stops
            and not _profanity.contains_profanity(token.lemma_)
        ]
        freq = Counter(lemmas)
    else:
        # Original fallback: regex strip → split → filter stop words
        cleaned = re.sub(r"[^a-zA-Z\s]", "", combined.lower())
        words = [w for w in cleaned.split()
                 if w not in STOP_WORDS and len(w) > 2
                 and not _profanity.contains_profanity(w)]
        freq = Counter(words)

    return [{"text": w, "value": c} for w, c in freq.most_common(80)]

def get_entries(q):
    conn = get_db()
    rows = conn.execute("SELECT * FROM sentiment_responses WHERE question_index=? ORDER BY id", (q,)).fetchall()
    conn.close()
    return [{"text": r["text"], "sentiment": {"polarity": r["polarity"], "subjectivity": r["subjectivity"],
             "label": r["label"], "emoji": r["emoji"], "color": r["color"]}, "timestamp": r["timestamp"]} for r in rows]

def stats_for(entries):
    if not entries:
        return {"positive": 0, "negative": 0, "neutral": 0, "avg_polarity": 0, "total": 0}
    p = sum(1 for e in entries if e["sentiment"]["label"] == "Positive")
    n = sum(1 for e in entries if e["sentiment"]["label"] == "Negative")
    u = sum(1 for e in entries if e["sentiment"]["label"] == "Neutral")
    a = sum(e["sentiment"]["polarity"] for e in entries) / len(entries)
    return {"positive": p, "negative": n, "neutral": u, "avg_polarity": round(a, 3), "total": len(entries)}


# ═══════════════════════════════════════════════════════
# TURING TEST ANALYTICS
# ═══════════════════════════════════════════════════════
def get_turing_stats(job_group=None):
    conn = get_db()
    jg = job_group

    # Total respondents (filtered)
    if jg:
        total_respondents = conn.execute(
            "SELECT COUNT(*) as c FROM turing_responses WHERE job_group=?", (jg,)).fetchone()["c"]
    else:
        total_respondents = conn.execute("SELECT COUNT(*) as c FROM turing_responses").fetchone()["c"]

    # By job group — always unfiltered so all bars show even when a filter is active
    by_job = {}
    for g in JOB_GROUPS:
        r = conn.execute("""SELECT COUNT(a.id) as total, SUM(a.correct) as cs
            FROM turing_answers a JOIN turing_responses r ON a.respondent_id=r.respondent_id
            WHERE r.job_group=?""", (g,)).fetchone()
        t, c = r["total"] or 0, r["cs"] or 0
        if t > 0:
            by_job[g] = {"total": t, "correct": c, "accuracy": round(c / t * 100, 1),
                         "fooled_pct": round((t - c) / t * 100, 1)}

    if total_respondents == 0:
        conn.close()
        return {"total_respondents": 0, "overall_accuracy": 0, "total_answers": 0, "total_correct": 0,
                "by_job_group": by_job, "by_seniority": {}, "by_item": {},
                "avg_ratings": {}, "trust_tasks": {}, "respondents_list": []}

    # Overall accuracy (filtered)
    if jg:
        r = conn.execute("""SELECT COUNT(a.id) as total, SUM(a.correct) as cs
            FROM turing_answers a JOIN turing_responses r ON a.respondent_id=r.respondent_id
            WHERE r.job_group=?""", (jg,)).fetchone()
        total_answers, total_correct = r["total"] or 0, r["cs"] or 0
    else:
        total_answers = conn.execute("SELECT COUNT(*) as c FROM turing_answers").fetchone()["c"]
        total_correct = conn.execute("SELECT COUNT(*) as c FROM turing_answers WHERE correct=1").fetchone()["c"]
    overall_accuracy = round(total_correct / total_answers * 100, 1) if total_answers else 0

    # By seniority (filtered)
    by_sen = {}
    for sl in SENIORITY_LEVELS:
        if jg:
            r = conn.execute("""SELECT COUNT(a.id) as total, SUM(a.correct) as cs
                FROM turing_answers a JOIN turing_responses r ON a.respondent_id=r.respondent_id
                WHERE r.seniority=? AND r.job_group=?""", (sl, jg)).fetchone()
        else:
            r = conn.execute("""SELECT COUNT(a.id) as total, SUM(a.correct) as cs
                FROM turing_answers a JOIN turing_responses r ON a.respondent_id=r.respondent_id
                WHERE r.seniority=?""", (sl,)).fetchone()
        t, c = r["total"] or 0, r["cs"] or 0
        if t > 0:
            by_sen[sl] = {"total": t, "correct": c, "accuracy": round(c / t * 100, 1),
                          "fooled_pct": round((t - c) / t * 100, 1)}

    # By item — per-item accuracy + avg ratings (filtered)
    by_item = {}
    for item in TURING_ITEMS:
        if jg:
            r = conn.execute("""SELECT COUNT(a.id) as total, SUM(a.correct) as cs,
                AVG(a.rating_trust) as rt, AVG(a.rating_empathy) as re,
                AVG(a.rating_safety) as rs, AVG(a.rating_usefulness) as ru
                FROM turing_answers a JOIN turing_responses rr ON a.respondent_id=rr.respondent_id
                WHERE a.item_id=? AND rr.job_group=?""", (item["id"], jg)).fetchone()
        else:
            r = conn.execute("""SELECT COUNT(*) as total, SUM(correct) as cs,
                AVG(rating_trust) as rt, AVG(rating_empathy) as re,
                AVG(rating_safety) as rs, AVG(rating_usefulness) as ru
                FROM turing_answers WHERE item_id=?""", (item["id"],)).fetchone()
        t, c = r["total"] or 0, r["cs"] or 0
        if t > 0:
            by_item[item["id"]] = {
                "section": item["section"],
                "label": item["label"],
                "total": t,
                "correct": c,
                "accuracy": round(c / t * 100, 1),
                "avg_ratings": {
                    "trust":       round(r["rt"] or 0, 2),
                    "empathy":     round(r["re"] or 0, 2),
                    "safety":      round(r["rs"] or 0, 2),
                    "usefulness":  round(r["ru"] or 0, 2),
                },
            }

    # Overall average ratings (filtered)
    if jg:
        rat = conn.execute("""SELECT AVG(a.rating_trust) as t, AVG(a.rating_empathy) as e,
                              AVG(a.rating_safety) as s, AVG(a.rating_usefulness) as u
                              FROM turing_answers a JOIN turing_responses r ON a.respondent_id=r.respondent_id
                              WHERE r.job_group=?""", (jg,)).fetchone()
    else:
        rat = conn.execute("""SELECT AVG(rating_trust) as t, AVG(rating_empathy) as e,
                              AVG(rating_safety) as s, AVG(rating_usefulness) as u
                              FROM turing_answers""").fetchone()
    avg_ratings = {"trust": round(rat["t"] or 0, 2), "empathy": round(rat["e"] or 0, 2),
                   "safety": round(rat["s"] or 0, 2), "usefulness": round(rat["u"] or 0, 2)}

    # Trust tasks (filtered)
    if jg:
        task_rows = conn.execute("""SELECT t.task, COUNT(*) as c FROM turing_tasks t
            JOIN turing_responses r ON t.respondent_id=r.respondent_id
            WHERE r.job_group=? GROUP BY t.task ORDER BY c DESC""", (jg,)).fetchall()
    else:
        task_rows = conn.execute("SELECT task, COUNT(*) as c FROM turing_tasks GROUP BY task ORDER BY c DESC").fetchall()
    trust_tasks = {r["task"]: r["c"] for r in task_rows}

    # Recent respondents (filtered)
    if jg:
        recent = conn.execute("SELECT * FROM turing_responses WHERE job_group=? ORDER BY id DESC LIMIT 20",
                              (jg,)).fetchall()
    else:
        recent = conn.execute("SELECT * FROM turing_responses ORDER BY id DESC LIMIT 20").fetchall()
    respondents_list = [{"respondent_id": r["respondent_id"], "job_group": r["job_group"],
                         "seniority": r["seniority"], "timestamp": r["timestamp"]} for r in recent]

    conn.close()
    return {
        "total_respondents": total_respondents,
        "overall_accuracy": overall_accuracy,
        "total_answers": total_answers,
        "total_correct": total_correct,
        "by_job_group": by_job,
        "by_seniority": by_sen,
        "by_item": by_item,
        "avg_ratings": avg_ratings,
        "trust_tasks": trust_tasks,
        "respondents_list": respondents_list,
    }


# ═══════════════════════════════════════════════════════
# ROUTES — PARTICIPANT LANDING & ADMIN
# ═══════════════════════════════════════════════════════
@app.route("/")
def landing():
    return render_template("landing.html")


@app.route("/admin")
def admin():
    return render_template("admin.html", questions=QUESTIONS, title=APP_TITLE,
                           stop_words=sorted(STOP_WORDS),
                           job_groups=JOB_GROUPS,
                           seniority_levels=SENIORITY_LEVELS,
                           trust_tasks=TRUST_TASKS)


@app.route("/api/submit", methods=["POST"])
def api_submit():
    d = request.json
    t = (d.get("text") or "").strip()
    q = d.get("question_index", 0)
    pid = (d.get("participant_id") or "").strip() or None
    if not t:
        return jsonify({"error": "Empty"}), 400
    if _profanity.contains_profanity(t):
        return jsonify({"error": "Inappropriate content"}), 400
    s = sentiment(t)
    conn = get_db()
    conn.execute("INSERT INTO sentiment_responses (question_index,text,polarity,subjectivity,label,emoji,color,timestamp,participant_id) VALUES (?,?,?,?,?,?,?,?,?)",
                 (q, t, s["polarity"], s["subjectivity"], s["label"], s["emoji"], s["color"], datetime.now().isoformat(), pid))
    conn.commit()
    conn.close()
    entry = {"text": t, "sentiment": s, "timestamp": datetime.now().isoformat()}
    return jsonify({"entry": entry})


@app.route("/api/q/<int:q>")
def api_q(q):
    entries = get_entries(q)
    texts = [e["text"] for e in entries]
    return jsonify({
        "words": extract_words(texts) if texts else [],
        "stats": stats_for(entries),
        "entries": entries[-15:]
    })


@app.route("/api/all")
def api_all():
    r = {}
    for i in range(len(QUESTIONS)):
        entries = get_entries(i)
        texts = [e["text"] for e in entries]
        r[str(i)] = {"words": extract_words(texts) if texts else [],
                      "stats": stats_for(entries), "entries": entries[-10:]}
    ae = []
    for i in range(len(QUESTIONS)):
        ae.extend(get_entries(i))
    at = [e["text"] for e in ae]
    r["agg"] = {"words": extract_words(at) if at else [], "stats": stats_for(ae)}
    # Include turing snapshot
    r["turing"] = get_turing_stats()
    return jsonify(r)


@app.route("/api/reset", methods=["POST"])
def api_reset():
    conn = get_db()
    conn.execute("DELETE FROM sentiment_responses")
    conn.execute("DELETE FROM turing_responses")
    conn.execute("DELETE FROM turing_answers")
    conn.execute("DELETE FROM turing_tasks")
    conn.execute("DELETE FROM acceptance_responses")
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/export")
def api_export():
    data = {}
    for i in range(len(QUESTIONS)):
        data[QUESTIONS[i]] = get_entries(i)
    data["turing_test"] = get_turing_stats()
    data["acceptance_survey"] = api_acceptance_stats().get_json()
    return jsonify(data)


@app.route("/api/export/excel")
def api_export_excel():
    wb = openpyxl.Workbook()
    bold = Font(bold=True)
    conn = get_db()

    # ── Sheet 1: Sentiment Responses — wide format, one row per participant ──
    ws1 = wb.active
    ws1.title = "Sentiment Responses"

    # Build wide headers: Participant ID + per-question columns
    sent_headers = ["Participant ID"]
    for i, q in enumerate(QUESTIONS):
        sent_headers += [f"Q{i+1} Transcript", f"Q{i+1} Sentiment", f"Q{i+1} Polarity"]
    sent_headers.append("Timestamp")
    ws1.append(sent_headers)
    for cell in ws1[1]: cell.font = bold

    # Group rows by participant_id; legacy rows (no ID) each get a unique placeholder
    sent_rows = conn.execute("SELECT * FROM sentiment_responses ORDER BY id").fetchall()
    groups = {}   # pid → {q_index: row}
    anon_n = 0
    for r in sent_rows:
        pid = r["participant_id"]
        if not pid:
            anon_n += 1
            pid = f"LEGACY-{anon_n:04d}"
        if pid not in groups:
            groups[pid] = {"_ts": r["timestamp"]}
        groups[pid][r["question_index"]] = r
        groups[pid]["_ts"] = r["timestamp"]   # track latest timestamp

    for pid, qs in groups.items():
        row = [pid]
        for i in range(len(QUESTIONS)):
            if i in qs:
                r = qs[i]
                row += [r["text"], r["label"], r["polarity"]]
            else:
                row += ["", "", ""]
        row.append(qs.get("_ts", ""))
        ws1.append(row)

    # ── Sheet 2: Turing Test — wide format, one row per respondent ───────────
    ws2 = wb.create_sheet("Turing Test")

    item_ids = [x["id"] for x in TURING_ITEMS]
    item_labels = {x["id"]: x["label"] for x in TURING_ITEMS}
    tt_headers = ["Participant ID", "Job Group", "Seniority", "Timestamp", "Trusted Tasks"]
    for iid in item_ids:
        lbl = item_labels[iid]
        tt_headers += [f"{lbl} — Correct", f"{lbl} — Trust",
                       f"{lbl} — Empathy", f"{lbl} — Safety", f"{lbl} — Usefulness"]
    ws2.append(tt_headers)
    for cell in ws2[1]: cell.font = bold

    tasks_map = {r["respondent_id"]: r["tasks"] for r in conn.execute(
        "SELECT respondent_id, GROUP_CONCAT(task, '; ') as tasks FROM turing_tasks GROUP BY respondent_id"
    ).fetchall()}

    ans_by_resp = {}
    for a in conn.execute("SELECT * FROM turing_answers ORDER BY id").fetchall():
        ans_by_resp.setdefault(a["respondent_id"], {})[a["item_id"]] = a

    for r in conn.execute("SELECT * FROM turing_responses ORDER BY id").fetchall():
        rid = r["respondent_id"]
        row = [rid, r["job_group"], r["seniority"], r["timestamp"], tasks_map.get(rid, "")]
        for iid in item_ids:
            a = ans_by_resp.get(rid, {}).get(iid)
            if a:
                row += ["Yes" if a["correct"] else "No",
                        a["rating_trust"], a["rating_empathy"], a["rating_safety"], a["rating_usefulness"]]
            else:
                row += ["", "", "", "", ""]
        ws2.append(row)

    # ── Sheet 3: AI Acceptance Survey — full question text as column headers ─
    ws3 = wb.create_sheet("AI Acceptance Survey")

    # Build (key, full_text) pairs in order
    likert_items = [(k, q_text)
                    for part_val in ACCEPTANCE_LIKERT.values()
                    for k, q_text in part_val["questions"].items()]
    open_refl_keys = [
        ("G1", "G1: Which tasks would you like AI to assist you with?"),
        ("G2", "G2: Biggest concern about AI in your area of work?"),
        ("G3", "G3: What would make you more confident using AI?"),
        ("G4", "G4: Anything else you'd like to share about AI in healthcare?"),
    ]

    ac_headers = ["Participant ID", "Timestamp", "Age Group", "Gender", "Cluster",
                  "Disciplines", "Years in Healthcare", "Years in Current Role", "Seniority",
                  "AI Usage Frequency", "AI Tools Used"]
    ac_headers += [text for _, text in likert_items]
    ac_headers += [label for _, label in open_refl_keys]
    ws3.append(ac_headers)
    for cell in ws3[1]: cell.font = bold

    for r in conn.execute("SELECT * FROM acceptance_responses ORDER BY id").fetchall():
        likert = json.loads(r["likert_answers"] or "{}")
        reflection = json.loads(r["open_reflection"] or "{}") if r["open_reflection"] else {}
        row = [r["participant_id"], r["timestamp"], r["age_group"], r["gender"],
               r["cluster"] or "",
               ", ".join(json.loads(r["disciplines"] or "[]")),
               r["years_healthcare"], r["years_role"], r["seniority"], r["ai_frequency"],
               ", ".join(json.loads(r["ai_tools"] or "[]"))]
        row += [likert.get(k, "") for k, _ in likert_items]
        row += [reflection.get(k, "") for k, _ in open_refl_keys]
        ws3.append(row)

    conn.close()

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"nuhs_summit_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=filename)


# ═══════════════════════════════════════════════════════
# ROUTES — TURING TEST
# ═══════════════════════════════════════════════════════
@app.route("/api/turing/submit", methods=["POST"])
def api_turing_submit():
    d = request.json
    rid = d.get("respondent_id") or str(uuid.uuid4())[:8]
    job_group = d.get("job_group", "")
    seniority = d.get("seniority", "")
    answers = d.get("answers", [])  # [{item_id, section, guess, ratings:{trust,empathy,safety,usefulness}}]
    tasks = d.get("tasks", [])

    if not answers:
        return jsonify({"error": "No answers"}), 400

    conn = get_db()
    ts = datetime.now().isoformat()
    conn.execute("INSERT INTO turing_responses (respondent_id,job_group,seniority,timestamp) VALUES (?,?,?,?)",
                 (rid, job_group, seniority, ts))

    results = []
    for a in answers:
        iid = a.get("item_id", "")
        section = a.get("section", "")
        guess = a.get("guess", "")
        item = next((x for x in TURING_ITEMS if x["id"] == iid), None)
        correct = 1 if (item and guess == item["correct_answer"]) else 0
        ratings = a.get("ratings", {})
        conn.execute("""INSERT INTO turing_answers
            (respondent_id,item_id,section,guess,correct,
             rating_trust,rating_empathy,rating_safety,rating_usefulness,timestamp)
            VALUES (?,?,?,?,?,?,?,?,?,?)""",
            (rid, iid, section, guess, correct,
             ratings.get("trust", 0), ratings.get("empathy", 0),
             ratings.get("safety", 0), ratings.get("usefulness", 0), ts))
        results.append({
            "item_id": iid,
            "correct": bool(correct),
            "correct_answer": item["correct_answer"] if item else "",
        })

    for task in tasks:
        conn.execute("INSERT INTO turing_tasks (respondent_id,task) VALUES (?,?)", (rid, task))

    conn.commit()
    conn.close()
    return jsonify({"ok": True, "respondent_id": rid, "results": results})


@app.route("/api/turing/stats")
def api_turing_stats():
    job_group = request.args.get("job_group") or None
    return jsonify(get_turing_stats(job_group=job_group))


@app.route("/api/turing/scenarios")
def api_turing_scenarios():
    """Return items without revealing correct answers (for external API consumers)."""
    safe = []
    for item in TURING_ITEMS:
        s = {"id": item["id"], "section": item["section"], "label": item["label"], "type": item["type"]}
        if item["type"] == "scenario":
            s["patient"] = item["patient"]
            s["response"] = item["response"]
        else:
            s["turns"] = item["turns"]
        safe.append(s)
    return jsonify({"items": safe, "job_groups": JOB_GROUPS,
                    "seniority_levels": SENIORITY_LEVELS, "trust_tasks": TRUST_TASKS})


# ═══════════════════════════════════════════════════════
# ROUTES — PARTICIPANT SURVEYS
# ═══════════════════════════════════════════════════════
@app.route("/survey")
def survey_redirect():
    return redirect(url_for("survey_turing"))


@app.route("/survey/turing")
def survey_turing():
    # is_ai and correct_answer are embedded for client-side inline reveal.
    # Security is not a concern — this is a fun booth activity, not an assessment.
    items_json = json.dumps([{
        "id": x["id"], "section": x["section"], "label": x["label"], "type": x["type"],
        "patient": x.get("patient", ""), "response": x.get("response", ""),
        "turns": x.get("turns", []),
        "is_ai": x["is_ai"], "correct_answer": x["correct_answer"],
    } for x in TURING_ITEMS])
    return render_template("survey_turing.html", title=APP_TITLE,
                           items_json=items_json,
                           job_groups=JOB_GROUPS, seniority_levels=SENIORITY_LEVELS,
                           trust_tasks=TRUST_TASKS)


@app.route("/survey/sentiment")
def survey_sentiment():
    return render_template("survey_sentiment.html", questions=QUESTIONS)


@app.route("/survey/acceptance")
def survey_acceptance():
    return render_template("survey_acceptance.html", part_a=ACCEPTANCE_PART_A, likert=ACCEPTANCE_LIKERT)


@app.route("/qr")
def qr_code():
    import qrcode
    host = request.host_url.rstrip("/")
    url = f"{host}/survey/turing"
    img = qrcode.make(url, box_size=10, border=2)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return send_file(buf, mimetype="image/png", download_name="survey_qr.png")


# ═══════════════════════════════════════════════════════
# ROUTES — AI ACCEPTANCE SURVEY
# ═══════════════════════════════════════════════════════
@app.route("/api/acceptance/submit", methods=["POST"])
def api_acceptance_submit():
    d = request.json or {}
    part_a = d.get("part_a", {})
    likert = d.get("likert_answers", {})

    # Validate required Part A single-select fields
    required = ["age_group", "gender", "cluster", "years_healthcare", "years_role", "ai_frequency"]
    missing = [f for f in required if not part_a.get(f)]
    if missing:
        return jsonify({"error": f"Missing required fields: {', '.join(missing)}"}), 400

    participant_id = str(uuid.uuid4())[:8]
    ts = datetime.now().isoformat()
    open_reflection = d.get("open_reflection", {})

    conn = get_db()
    conn.execute(
        """INSERT INTO acceptance_responses
           (participant_id, timestamp, age_group, gender, cluster, disciplines,
            years_healthcare, years_role, seniority, ai_frequency, ai_tools,
            likert_answers, open_reflection)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (
            participant_id, ts,
            part_a.get("age_group", ""),
            part_a.get("gender", ""),
            part_a.get("cluster", ""),
            json.dumps(part_a.get("disciplines", [])),
            part_a.get("years_healthcare", ""),
            part_a.get("years_role", ""),
            part_a.get("seniority", ""),
            part_a.get("ai_frequency", ""),
            json.dumps(part_a.get("ai_tools", [])),
            json.dumps(likert),
            json.dumps(open_reflection),
        )
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "participant_id": participant_id})


@app.route("/api/acceptance/stats")
def api_acceptance_stats():
    conn = get_db()
    rows = conn.execute("SELECT * FROM acceptance_responses ORDER BY id").fetchall()
    conn.close()

    total = len(rows)
    if total == 0:
        return jsonify({
            "total_respondents": 0,
            "part_a": {},
            "likert_averages": {},
        })

    # Part A — frequency distributions for each categorical field
    part_a_fields = ["age_group", "gender", "cluster", "disciplines", "years_healthcare",
                     "years_role", "seniority", "ai_frequency", "ai_tools"]
    part_a_stats = {}
    for field in part_a_fields:
        counts = {}
        for row in rows:
            val = row[field] or ""
            # disciplines and ai_tools are JSON lists
            if field in ("disciplines", "ai_tools"):
                items = json.loads(val) if val else []
                for item in items:
                    counts[item] = counts.get(item, 0) + 1
            else:
                counts[val] = counts.get(val, 0) + 1
        part_a_stats[field] = counts

    # Likert — average score per question key
    sums = {}
    counts_l = {}
    for row in rows:
        answers = json.loads(row["likert_answers"]) if row["likert_answers"] else {}
        for key, val in answers.items():
            try:
                v = int(val)
                sums[key] = sums.get(key, 0) + v
                counts_l[key] = counts_l.get(key, 0) + 1
            except (ValueError, TypeError):
                pass

    likert_averages = {}
    for part_key, part_data in ACCEPTANCE_LIKERT.items():
        part_avgs = {}
        for q_key, q_text in part_data["questions"].items():
            n = counts_l.get(q_key, 0)
            avg = round(sums.get(q_key, 0) / n, 2) if n > 0 else None
            part_avgs[q_key] = {"question": q_text, "avg": avg, "n": n}
        likert_averages[part_key] = {
            "title": part_data["title"],
            "questions": part_avgs,
        }

    return jsonify({
        "total_respondents": total,
        "part_a": part_a_stats,
        "likert_averages": likert_averages,
    })


if __name__ == "__main__":
    print("=" * 60)
    print(f"  {APP_TITLE}  (v4 — with Turing Test)")
    print("=" * 60)
    print(f"  Landing page:  http://localhost:5001/")
    print(f"  Admin:         http://localhost:5001/admin")
    print(f"  Survey/Turing: http://localhost:5001/survey/turing")
    print(f"  QR Code image: http://localhost:5001/qr")
    print(f"  Questions:     {len(QUESTIONS)}")
    print(f"  Scenarios:     {len(SCENARIOS)}")
    print(f"  Database:      {DB_PATH} (SQLite, persists)")
    print("=" * 60)
    port = int(os.environ.get("PORT", 5001))
    debug = os.environ.get("FLASK_DEBUG", "false").lower() == "true"
    app.run(host="0.0.0.0", port=port, debug=debug)
